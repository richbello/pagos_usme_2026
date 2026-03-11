import os, re, glob, pdfplumber, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

CARPETA_PDF   = r"C:\RICHARD\FDL\Usme\2026\Pagos\Marzo\Lote_1"
CARPETA_EXCEL = r"C:\RICHARD\FDL\Usme\2026\Pagos\Marzo\Lote_1"
NOMBRE_EXCEL  = "Consolidado_Pagos_Lote1.xlsx"

COLUMNAS = [
    "Archivo PDF","Ciudad y Fecha","Documento No.","Nombre Contratista",
    "Entidad (Debe A)","NIT Entidad","Cedula Contratista","Por Concepto",
    "Periodo Desde","Periodo Hasta","Fecha Suspension Inicio","Fecha Suspension Fin",
    "La Suma De ($)","No. Contrato","Tipo Contrato","Planillas Seg. Social",
    "Periodo de Pago","Riesgo","Fecha de Pago","Banco / Entidad Financiera",
    "Tipo de Cuenta","No. de Cuenta","Ingreso Base Cotizacion ($)",
    "Pago No.","Total Pagos","Direccion","Telefono",
]

def limpiar(t):
    return re.sub(r'\s+', ' ', str(t)).strip() if t else ""

def buscar(texto, patron, grupo=1, flags=re.IGNORECASE|re.DOTALL):
    try:
        m = re.search(patron, texto, flags)
        return limpiar(m.group(grupo)) if m else ""
    except:
        return ""

def extraer_datos(ruta_pdf):
    datos = {col: "" for col in COLUMNAS}
    datos["Archivo PDF"] = os.path.basename(ruta_pdf)
    try:
        with pdfplumber.open(ruta_pdf) as pdf:
            # Solo extraemos la pagina que tiene el comprobante de cobro
            # Buscamos la pagina que tenga "Favor consignar en" y "DEBE A"
            texto_comprobante = ""
            texto_completo = ""
            for pagina in pdf.pages:
                t = pagina.extract_text() or ""
                texto_completo += t + "\n"
                if "Favor consignar en" in t or "DEBE A:" in t or "La suma de" in t:
                    texto_comprobante += t + "\n"

            # Usamos el comprobante si lo encontramos, si no el texto completo
            texto = texto_comprobante if texto_comprobante.strip() else texto_completo

        if not texto_completo.strip():
            print(f"  SIN TEXTO: {datos['Archivo PDF']}")
            return datos

        # ── Ciudad y Fecha ─────────────────────────────────────────
        datos["Ciudad y Fecha"] = buscar(
            texto, r'Ciudad\s+y\s+fecha[:\s]+(.+?)(?=Documento\s+No\.)'
        )

        # ── Documento No. ──────────────────────────────────────────
        datos["Documento No."] = buscar(
            texto, r'Documento\s+No\.?\s*([\d\.\s]+?)(?=\n|Yo\b)'
        )
        if not datos["Documento No."]:
            datos["Documento No."] = buscar(
                texto, r'Documento\s+No\.?\s*([\d\.]+)'
            )

        # ── Nombre Contratista ─────────────────────────────────────
        datos["Nombre Contratista"] = buscar(
            texto, r'\bYo\b\s+([\w\s]+?)\s+identificado'
        )
        if not datos["Nombre Contratista"]:
            datos["Nombre Contratista"] = buscar(
                texto, r'DEBE\s+A[:\s]*\n?\s*([\w\s]+?)\s*,'
            )

        # ── Entidad ────────────────────────────────────────────────
        datos["Entidad (Debe A)"] = buscar(
            texto, r'((?:EL\s+)?FONDO\s+DE\s+DESARROLLO\s+LOCAL\s+DE\s+\w+)'
        )

        # ── NIT ────────────────────────────────────────────────────
        datos["NIT Entidad"] = buscar(
            texto, r'NIT\s*([\d\.\-]+)'
        )

        # ── Cedula Contratista ─────────────────────────────────────
        datos["Cedula Contratista"] = buscar(
            texto, r'c[eé]dula\s+de\s+ciudadan[ií]a\s+([\d\.\s]+?)(?=\n|PRESTAR|APOYAR|Por)'
        )
        if not datos["Cedula Contratista"]:
            datos["Cedula Contratista"] = buscar(
                texto, r'C\.C\.?\s*No\.?\s*([\d\.]+)'
            )

        # ── Por Concepto ───────────────────────────────────────────
        datos["Por Concepto"] = buscar(
            texto, r'Por\s+concepto[:\s]*(.*?)(?=Periodo[:\s]|\nPer|\nFecha)'
        )

        # ── Periodo Desde / Hasta ──────────────────────────────────
        m = re.search(
            r'Periodo[:\s_]+([\d]+\s+de\s+\w+\s+de\s+\d{4})\s+a\s+([\d]+\s+de\s+\w+\s+de\s+\d{4})',
            texto, re.IGNORECASE
        )
        if m:
            datos["Periodo Desde"] = limpiar(m.group(1))
            datos["Periodo Hasta"] = limpiar(m.group(2))

        # ── Fecha Suspension ───────────────────────────────────────
        m = re.search(
            r'suspensi[o\u00f3]n[^:]*:?\s*(N\s*/\s*A|N/A|[\d/\-]+)\s+a\s+(N\s*/\s*A|N/A|[\d/\-]+)',
            texto, re.IGNORECASE
        )
        if m:
            datos["Fecha Suspension Inicio"] = limpiar(m.group(1))
            datos["Fecha Suspension Fin"]    = limpiar(m.group(2))

        # ── La Suma De ─────────────────────────────────────────────
        datos["La Suma De ($)"] = buscar(
            texto, r'La\s+suma\s+de[:\s]*\$?\s*([\d\.,]+)'
        )

        # ── No. Contrato ───────────────────────────────────────────
        # Busca patron tipo 0624-2025 o 0520-2025
        datos["No. Contrato"] = buscar(
            texto, r'No\.?\s*Contrato[:\s]*([\d]{3,4}[\-/][\d]{4})'
        )
        if not datos["No. Contrato"]:
            # Busca el numero al inicio del documento (ej: 0624-2025 solo en primera linea)
            datos["No. Contrato"] = buscar(
                texto, r'^([\d]{3,4}[\-/][\d]{4})', flags=re.MULTILINE
            )
        if not datos["No. Contrato"]:
            datos["No. Contrato"] = buscar(
                texto, r'Contrato[:\s#Nno\.]*\s*([\d]{3,4}[\-/][\d]{4})'
            )

        # ── Tipo Contrato ──────────────────────────────────────────
        datos["Tipo Contrato"] = buscar(
            texto, r'Tipo\s+contrato[:\s]*([\w\s]+?)(?=Planillas|\n|$)'
        )

        # ── Planillas Seg. Social ──────────────────────────────────
        datos["Planillas Seg. Social"] = buscar(
            texto, r'Planillas?\s+pago\s+seguridad\s+social[:\s]*([\d]+)'
        )

        # ── Periodo de Pago ────────────────────────────────────────
        datos["Periodo de Pago"] = buscar(
            texto, r'Periodo\s+de\s+pago[:\s]*([\w]+)'
        )

        # ── Riesgo ─────────────────────────────────────────────────
        datos["Riesgo"] = buscar(
            texto, r'Riesgo[:\s]*(\d+)'
        )

        # ── Fecha de Pago ──────────────────────────────────────────
        datos["Fecha de Pago"] = buscar(
            texto, r'Fecha\s+de\s+pago[:\s]*([\d]+\s+de\s+\w+\s+de\s+\d{4})'
        )
        if not datos["Fecha de Pago"]:
            # Formato: "Febrero 12 de 2026"
            datos["Fecha de Pago"] = buscar(
                texto, r'Fecha\s+de\s+pago[:\s]*(\w+\s+\d+\s+de\s+\d{4})'
            )
        if not datos["Fecha de Pago"]:
            datos["Fecha de Pago"] = buscar(
                texto, r'Fecha\s+de\s+pago[:\s]*([\d/\-]+)'
            )

        # ── Banco / Entidad Financiera ─────────────────────────────
        # CORREGIDO: acepta bancos cortos como "NU" (2 letras)
        # y nombres con caracteres especiales como BOGOTA con tilde
        m = re.search(
            r'consignar\s+en[:\s]+(.+?)\s+Cuenta[:\s]',
            texto, re.IGNORECASE
        )
        if m:
            banco = limpiar(m.group(1))
            # Filtra si captura texto muy largo o sin sentido
            if len(banco) > 0 and len(banco) < 60:
                datos["Banco / Entidad Financiera"] = banco

        # ── Tipo de Cuenta ─────────────────────────────────────────
        # CORREGIDO: acepta AHORRO o AHORROS
        datos["Tipo de Cuenta"] = buscar(
            texto, r'Cuenta[:\s]+(AHORROS?|CORRIENTE)', flags=re.IGNORECASE
        )

        # ── No. de Cuenta ──────────────────────────────────────────
        # CORREGIDO: busca el numero despues de AHORRO/AHORROS/CORRIENTE + N°
        # El simbolo grado puede ser °, º, o el caracter unicode
        datos["No. de Cuenta"] = buscar(
            texto,
            r'(?:AHORROS?|CORRIENTE|AHORRO)\s+N[o\u00b0\u00ba\u00b0\s°\u02da]?\.?\s*([\d\-]{6,20})',
            flags=re.IGNORECASE
        )
        if not datos["No. de Cuenta"]:
            # Alternativa: N° seguido del numero en la misma zona de "consignar en"
            datos["No. de Cuenta"] = buscar(
                texto,
                r'N[o°\u00b0\u00ba]\.?\s+([\d][\d\-]{5,19})',
                flags=re.IGNORECASE
            )
        if not datos["No. de Cuenta"]:
            # Ultima alternativa: numero de 7+ digitos cerca de Cuenta
            datos["No. de Cuenta"] = buscar(
                texto,
                r'Cuenta[^\d]+([\d]{7,20})',
                flags=re.IGNORECASE
            )

        # ── Ingreso Base de Cotizacion ─────────────────────────────
        datos["Ingreso Base Cotizacion ($)"] = buscar(
            texto,
            r'Ingreso\s+base\s+de\s+cotizaci[o\u00f3]n[:\s]*\$?\s*([\d\.,]+)'
        )

        # ── Pago X de Y ────────────────────────────────────────────
        m = re.search(r'Pago[:\s]+(\d+)\s+de\s+(\d+)', texto, re.IGNORECASE)
        if m:
            datos["Pago No."]    = limpiar(m.group(1))
            datos["Total Pagos"] = limpiar(m.group(2))

        # ── Direccion ──────────────────────────────────────────────
        datos["Direccion"] = buscar(
            texto,
            r'DIRECCI[O\u00d3]N[:\s]*([\w\s\d#\-]+?)(?=TEL[E\u00c9]FONO|TELEFONO|$)'
        )

        # ── Telefono ───────────────────────────────────────────────
        datos["Telefono"] = buscar(
            texto, r'TEL[E\u00c9]FONO[:\s]*([\d\s]+)'
        )

    except Exception as e:
        print(f"  ERROR en {datos['Archivo PDF']}: {e}")

    return datos


def crear_excel(lista_datos, ruta_salida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidado Pagos"

    f_tit = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    f_enc = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    f_dat = Font(name="Calibri", size=9)
    r_tit = PatternFill("solid", fgColor="0D3E6B")
    r_enc = PatternFill("solid", fgColor="1F4E79")
    r_par = PatternFill("solid", fgColor="DDEEFF")
    r_imp = PatternFill("solid", fgColor="FFFFFF")
    a_cen = Alignment(horizontal="center", vertical="center", wrap_text=True)
    a_izq = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNAS))
    c = ws.cell(row=1, column=1,
        value="CONSOLIDADO COMPROBANTES DE PAGO - FDL USME 2026 Lote 1 | Generado: " +
              datetime.now().strftime("%d/%m/%Y %H:%M"))
    c.font = f_tit; c.fill = r_tit; c.alignment = a_cen
    ws.row_dimensions[1].height = 26

    for idx, nombre in enumerate(COLUMNAS, 1):
        c = ws.cell(row=2, column=idx, value=nombre)
        c.font = f_enc; c.fill = r_enc; c.alignment = a_cen; c.border = borde
    ws.row_dimensions[2].height = 42

    for fi, datos in enumerate(lista_datos, 3):
        relleno = r_par if fi % 2 == 0 else r_imp
        for ci, col in enumerate(COLUMNAS, 1):
            c = ws.cell(row=fi, column=ci, value=datos.get(col, ""))
            c.font = f_dat; c.fill = relleno; c.alignment = a_izq; c.border = borde
        ws.row_dimensions[fi].height = 18

    anchos = {
        "Archivo PDF": 28, "Ciudad y Fecha": 26, "Documento No.": 16,
        "Nombre Contratista": 34, "Entidad (Debe A)": 38, "NIT Entidad": 16,
        "Cedula Contratista": 20, "Por Concepto": 58, "Periodo Desde": 22,
        "Periodo Hasta": 22, "Fecha Suspension Inicio": 22, "Fecha Suspension Fin": 22,
        "La Suma De ($)": 18, "No. Contrato": 18, "Tipo Contrato": 24,
        "Planillas Seg. Social": 22, "Periodo de Pago": 18, "Riesgo": 10,
        "Fecha de Pago": 24, "Banco / Entidad Financiera": 28, "Tipo de Cuenta": 16,
        "No. de Cuenta": 20, "Ingreso Base Cotizacion ($)": 24,
        "Pago No.": 12, "Total Pagos": 14, "Direccion": 30, "Telefono": 16,
    }
    for idx, nombre in enumerate(COLUMNAS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = anchos.get(nombre, 18)

    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNAS))}{len(lista_datos) + 2}"
    wb.save(ruta_salida)
    print(f"\n  Excel guardado en: {ruta_salida}")


def main():
    print("=" * 60)
    print("   EXTRACTOR COMPROBANTES DE PAGO - FDL USME 2026")
    print("=" * 60)
    archivos = sorted(glob.glob(os.path.join(CARPETA_PDF, "*.pdf")))
    if not archivos:
        print(f"\n  No se encontraron PDFs en:\n  {CARPETA_PDF}")
        return
    print(f"\n  PDFs encontrados: {len(archivos)}\n")
    todos = []; sin_datos = 0
    for i, ruta in enumerate(archivos, 1):
        nombre = os.path.basename(ruta)
        print(f"  [{i:03d}/{len(archivos)}] {nombre}", end=" ... ")
        d = extraer_datos(ruta); todos.append(d)
        tiene = any(v for k, v in d.items() if k != "Archivo PDF")
        print("OK" if tiene else "SIN DATOS")
        if not tiene: sin_datos += 1
    crear_excel(todos, os.path.join(CARPETA_EXCEL, NOMBRE_EXCEL))
    print("\n" + "=" * 60)
    print(f"  Procesados : {len(archivos)}")
    print(f"  Con datos  : {len(archivos) - sin_datos}")
    print(f"  Sin datos  : {sin_datos}")
    print("=" * 60 + "\n")

if __name__ == "__main__":
    main()
