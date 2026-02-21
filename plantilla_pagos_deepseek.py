import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

def procesar_pagos_consolidado():
    # Ruta de entrada y salida
    ruta_entrada = r"C:\RICHARD\FDL\Usme\2026\Pagos\Febrero\Extracción-grupo3_feb.xlsx"
    ruta_destino = r"C:\RICHARD\FDL\Usme\2026\Pagos\Febrero\PLANTILLA_PAGOS_GENERADAFEB.xlsx"
    
    # Obtener fecha actual en formato YYYYMMDD
    fecha_actual = datetime.now().strftime("%Y%m%d")
    print(f"📅 Fecha actual para columnas C y F: {fecha_actual}")
    
    # Leer el archivo consolidado
    print(f"Leyendo archivo: {ruta_entrada}")
    try:
        df = pd.read_excel(ruta_entrada)
        print(f"✓ Archivo leído: {len(df)} filas")
        
        # Mostrar las columnas que realmente tiene el archivo
        print(f"\nColumnas en el archivo:")
        for i, col in enumerate(df.columns, 1):
            print(f"  {i:2d}. {col}")
        
    except Exception as e:
        print(f"✗ Error al leer: {e}")
        return
    
    # ===== TABLA DE EQUIVALENCIAS RETECA % -> INDICADOR =====
    # Esto es exactamente lo que me diste, sin cambios
    equivalencias = {
        "0,100%": "01",
        "0,050%": "02",
        "0,200%": "03",
        "0,100%": "05",
        "0,110%": "06",
        "0,050%": "07",
        "2,000%": "08",
        "2,000%": "09",
        "0,350%": "10",
        "0,400%": "11",
        "1,000%": "12",
        "0,010%": "13",
        "0,100%": "14",
        "0,150%": "15",
        "0,250%": "16",
        "0,350%": "17",
        "0,600%": "18",
        "0,200%": "19",
        "0,250%": "20",
        "1,000%": "21",
        "1,100%": "22",
        "0,350%": "23",
        "0,600%": "24",
        "0,700%": "26",
        "0,400%": "27",
        "0,100%": "28",
        "0,200%": "29",
        "0,350%": "30",
        "0,400%": "31",
        "0,600%": "32",
        "1,104%": "33",
        "1,380%": "34",
        "0,414%": "35",
        "0,690%": "36",
        "0,700%": "37",
        "0,800%": "38",
        "0,966%": "39",
        "1,500%": "40",
        "0,250%": "41",
        "0,500%": "42",
        "0,712%": "86",
        "0,766%": "87",
        "0,866%": "88",
        "0,998%": "89",
        "1,014%": "91",
        "1,200%": "92",
        "1,214%": "R5",
        "1,400%": "94",
        "0,760%": "R3",
        "0,736%": "96",
        "1,030%": "97",
        "1,062%": "R4",
        "1,176%": "98",
        "1,254%": "99"
    }
    
    # Preparar columna de % Reteica para mapeo del indicador
    # La extracción guarda el % en "Pct_Reteica" o en "Valor Reteica" (ej: "0,966" o "0,966%")
    col_pct = None
    for c in ['Pct_Reteica', 'Valor Reteica']:
        if c in df.columns:
            col_pct = c
            break
    if col_pct is None:
        for c in df.columns:
            if 'reteica' in str(c).lower():
                col_pct = c
                break

    if col_pct:
        # Normalizar: asegurar formato "0,966%" para que haga match en equivalencias
        def normalizar_pct(val):
            s = str(val).strip().replace('.', ',')
            if not s.endswith('%'):
                s = s + '%'
            return s
        df["Reteica %"] = df[col_pct].apply(normalizar_pct)
        print(f"✓ Usando columna '{col_pct}' para indicador. Ejemplos: {df['Reteica %'].head(3).tolist()}")
    else:
        print("✗ No se encontró columna de % reteica")
        df["Reteica %"] = ""
    
    # Mapear cada valor de "Reteica %" al indicador correspondiente
    df["Indicador_Calculado"] = df["Reteica %"].map(equivalencias)
    
    # Mostrar algunos ejemplos del mapeo
    print(f"\n📊 Ejemplos de mapeo Reteica % -> Indicador:")
    for i in range(min(5, len(df))):
        reteica_val = df.iloc[i]["Reteica %"]
        indicador_val = df.iloc[i]["Indicador_Calculado"]
        print(f"  Pago {i+1}: '{reteica_val}' -> '{indicador_val}'")
    
    # Contar cuántos valores se mapearon correctamente
    mapeados = df["Indicador_Calculado"].notna().sum()
    print(f"✓ {mapeados}/{len(df)} valores mapeados a indicadores")
    
    # Crear archivo Excel de salida
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    
    # Encabezados exactos de la plantilla
    headers = [
        'Tipo Registro P', 'Clave Contab.', 'Codigo de la cuenta', 'Tipo Ident',
        'No Identificación', 'Indicador CME', 'Cuenta contable', 'importe',
        'Indicador de IVA', 'RP Doc Presupuestal', 'Posc Doc Pres', 'Pros Pre',
        'Programa de financiación', 'Fondo', 'Centro Gestor', 'Centro de costo',
        'Centro Beneficio', 'Orden CO', 'Elemento PEP', 'Grafo', 'Area funcional',
        'Segmento', 'Fecha Base', 'Condicion de Pago', 'Asignación', 'Texto',
        'Bloqueo Pago', 'Receptor Alternativo', 'Tipo Ident', 'No Identificación',
        'Via de Pago', 'Banco Propio', 'Id Cta', 'Ref 1', 'Ref 2', 'Referencia Pago',
        'Código Bco', 'No Cuenta', 'Tipo Cta', 'Tipo de retenciones',
        'Indicador de retención', 'Base imponible de retención', 'Importe de retención'
    ]
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        ws.cell(row=1, column=col_num).font = Font(bold=True)
    
    # Iniciar en fila 2
    fila_actual = 2
    
    # Procesar cada pago del consolidado
    for idx, row in df.iterrows():
        pago_num = idx + 1
        print(f"\n--- Procesando Pago {pago_num} ---")
        
        # ===== BUSCAR TODOS LOS DATOS EN EL CONSOLIDADO =====
        # 1. Buscar No Identificación (columna E en filas P31)
        no_identificacion = ""
        
        # Buscar primero por nombre exacto NIT_CC, luego por variantes
        # IMPORTANTE: NO incluir 'documento' para evitar match con "Documento No."
        for col in df.columns:
            col_strip = str(col).strip().upper()
            col_lower = str(col).lower()
            # Prioridad 1: columna exacta NIT_CC o equivalente
            if col_strip in ['NIT_CC', 'NIT/CC', 'NIT', 'CC']:
                valor = row[col]
                if pd.notna(valor):
                    no_identificacion = str(valor).strip()
                    print(f"✓ No Identificación encontrado en '{col}': {no_identificacion}")
                    break
            # Prioridad 2: columnas con nit_cc o cedula (NO 'documento', NO 'id' suelto)
            elif any(word in col_lower for word in ['nit_cc', 'cedula', 'identificacion']):
                valor = row[col]
                if pd.notna(valor):
                    no_identificacion = str(valor).strip()
                    print(f"✓ No Identificación encontrado en '{col}': {no_identificacion}")
                    break
        
        # Si no se encontró, usar valor por defecto
        if not no_identificacion:
            no_identificacion = f"ID{pago_num:04d}"
            print(f"⚠ No se encontró No Identificación, usando: {no_identificacion}")
        
        # 2. Buscar Valor Bruto
        valor_bruto = 0
        for col in df.columns:
            if 'valor' in str(col).lower() and 'bruto' in str(col).lower():
                valor_bruto = row[col] if pd.notna(row[col]) else 0
                print(f"✓ Valor Bruto encontrado en '{col}': {valor_bruto}")
                break
        
        # 3. Base imponible de retención — columna "BASE RETEICA" (limpiar $ y espacios)
        base_retencion = 0
        for col in df.columns:
            if str(col).strip().upper() == 'BASE RETEICA':
                val = row[col]
                if pd.notna(val):
                    # Limpiar: quitar $, espacios, puntos de miles → float
                    limpio = re.sub(r'[^\d,]', '', str(val)).replace(',', '.')
                    try:
                        base_retencion = float(limpio)
                    except:
                        base_retencion = 0
                print(f"✓ Base imponible: {base_retencion}")
                break

        # 4. Importe de retención — columna "TOTAL DESCUENTOS" (monto retenido real en pesos)
        importe_retencion = 0
        for col in df.columns:
            if str(col).strip().upper() == 'TOTAL DESCUENTOS':
                val = row[col]
                if pd.notna(val):
                    try:
                        importe_retencion = float(val)
                    except:
                        importe_retencion = 0
                print(f"✓ Importe de retención: {importe_retencion}")
                break
        
        # 5. Buscar RP Doc Presupuestal (para columna J en filas P40)
        # IMPORTANTE: NO usar 'doc' suelto — evita match con "Documento No."
        rp_doc = ""
        for col in df.columns:
            col_strip = str(col).strip().upper()
            col_lower = str(col).lower()
            if col_strip in ['RP DOC', 'RP DOC PRESUPUESTAL', 'RP_DOC', 'PRESUPUESTAL']:
                valor = row[col]
                if pd.notna(valor):
                    rp_doc = str(valor).strip()
                    print(f"✓ RP Doc Presupuestal encontrado en '{col}': {rp_doc}")
                    break
            elif 'presupuestal' in col_lower or ('rp' in col_lower and 'doc' in col_lower):
                valor = row[col]
                if pd.notna(valor):
                    rp_doc = str(valor).strip()
                    print(f"✓ RP Doc Presupuestal encontrado en '{col}': {rp_doc}")
                    break
        
        # Si no se encontró RP Doc, usar valor por defecto
        if not rp_doc:
            rp_doc = f"50009973{pago_num:02d}"
            print(f"⚠ No se encontró RP Doc Presupuestal, usando: {rp_doc}")
        
        # 6. Extraer asignación (número de contrato)
        asignacion = ""
        for col in df.columns:
            if 'contrato' in str(col).lower():
                contrato = str(row[col]).strip() if pd.notna(row[col]) else ""
                # Extraer números del contrato
                numeros = re.findall(r'\d+', contrato)
                if numeros:
                    if len(numeros) >= 2:
                        asignacion = f"{numeros[0]}-{numeros[1]}"
                    else:
                        asignacion = numeros[0]
                print(f"✓ Contrato encontrado en '{col}': {contrato}")
                print(f"✓ Asignación generada: {asignacion}")
                break
        
        if not asignacion:
            asignacion = f"{pago_num:03d}-2025"
            print(f"⚠ Sin contrato, usando asignación por defecto: {asignacion}")
        
        # 7. Buscar Código Banco
        codigo_bco = ""
        for col in df.columns:
            col_lower = str(col).lower()
            if any(word in col_lower for word in ['código', 'codigo']) and 'bco' in col_lower:
                valor = row[col]
                if pd.notna(valor):
                    codigo_bco = str(valor).strip()
                    print(f"✓ Código Banco encontrado en '{col}': {codigo_bco}")
                    break
        
        if not codigo_bco:
            codigo_bco = "051"
            print(f"⚠ No se encontró Código Banco, usando: {codigo_bco}")
        
        # 8. Buscar No Cuenta
        no_cuenta = ""
        for col in df.columns:
            col_lower = str(col).lower()
            if 'no' in col_lower and 'cuenta' in col_lower:
                valor = row[col]
                if pd.notna(valor):
                    no_cuenta = str(valor).strip()
                    print(f"✓ No Cuenta encontrado en '{col}': {no_cuenta}")
                    break
        
        if not no_cuenta:
            no_cuenta = "0550488435468647"
            print(f"⚠ No se encontró No Cuenta, usando: {no_cuenta}")
        
        # 9. Buscar Tipo Cta
        tipo_cta = ""
        for col in df.columns:
            col_lower = str(col).lower()
            if 'tipo' in col_lower and 'cta' in col_lower:
                valor = row[col]
                if pd.notna(valor):
                    tipo_cta = str(valor).strip()
                    print(f"✓ Tipo Cta encontrado en '{col}': {tipo_cta}")
                    break
        
        if not tipo_cta:
            tipo_cta = "02"
            print(f"⚠ No se encontró Tipo Cta, usando: {tipo_cta}")
        
        print(f"✓ Valor Bruto: {valor_bruto}")
        print(f"✓ RP Doc Presupuestal: {rp_doc}")
        print(f"✓ Base Retención (AP): {base_retencion}")
        print(f"✓ Importe Retención (AQ): {importe_retencion}")
        
        # ===== OBTENER INDICADOR DE RETECA % =====
        # Esta es la parte NUEVA que agregamos según tu requerimiento
        indicador_retencion = df.iloc[idx]["Indicador_Calculado"]
        
        # Si no se encontró indicador, usar el valor por defecto "39" como estaba antes
        if pd.isna(indicador_retencion) or not indicador_retencion:
            indicador_retencion = "39"
            print(f"⚠ No se encontró indicador para Reteica %, usando por defecto: {indicador_retencion}")
        else:
            print(f"✓ Indicador obtenido de Reteica %: {indicador_retencion}")
        
        # ===== TEXTO COLUMNA Z: "Pago No. 7 del 01/12/2025 al 31/12/2025" =====
        del_val, al_val, pago_no_real = "", "", ""
        for col in df.columns:
            col_s = str(col).strip().upper()
            val   = row[col]
            if col_s == "DEL" and pd.notna(val):
                del_val = str(val).strip()
            elif col_s == "AL" and pd.notna(val):
                al_val = str(val).strip()
            elif col_s == "PAGO NO." and pd.notna(val):
                try:
                    pago_no_real = str(int(val))
                except:
                    pago_no_real = str(val).strip()
        texto_z = f"Pago No. {pago_no_real} del {del_val} al {al_val}".strip()
        print(f"✓ Texto columna Z: '{texto_z}'")

        # ===== FILA C (PRIMERA FILA DEL BLOQUE) =====
        # Fila 2, 5, 8, 11... - COLUMNA E = "1001", COLUMNA C y F = fecha actual
        ws.cell(row=fila_actual, column=1, value='C')  # A: Tipo Registro P
        ws.cell(row=fila_actual, column=2, value=pago_num)  # B: Clave Contab.
        ws.cell(row=fila_actual, column=3, value=fecha_actual)  # C: Codigo de la cuenta = FECHA ACTUAL
        ws.cell(row=fila_actual, column=4, value='KR')  # D: Tipo Ident = "KR"
        ws.cell(row=fila_actual, column=5, value='1001')  # E: No Identificación = "1001"
        ws.cell(row=fila_actual, column=6, value=fecha_actual)  # F: Indicador CME = FECHA ACTUAL
        ws.cell(row=fila_actual, column=7, value='')  # G: Cuenta contable
        ws.cell(row=fila_actual, column=8, value='COP')  # H: importe (texto)
        ws.cell(row=fila_actual, column=10, value=asignacion)  # J: RP Doc Presupuestal
        
        # Nombre del contratista (columna K)
        nombre_contratista = ""
        for col in df.columns:
            if 'contratista' in str(col).lower():
                nombre = str(row[col]).strip() if pd.notna(row[col]) else ""
                # Limpiar NIT/C.C. del nombre
                nombre_limpio = re.sub(r'\s*(?:NIT\.|C\.C\.)\s*[\d\.,\s]+$', '', nombre).strip()
                nombre_contratista = nombre_limpio or f"CONTRATISTA {pago_num}"
                print(f"✓ Contratista encontrado: {nombre_contratista[:50]}...")
                break
        
        ws.cell(row=fila_actual, column=11, value=nombre_contratista)  # K: Posc Doc Pres
        
        # ===== FILA P40 (SEGUNDA FILA DEL BLOQUE) =====
        # Fila 3, 6, 9, 12... - COLUMNA E = VACÍO, COLUMNA J = RP Doc Presupuestal
        ws.cell(row=fila_actual+1, column=1, value='P')  # A: Tipo Registro P
        ws.cell(row=fila_actual+1, column=2, value=40)  # B: Clave Contab.
        ws.cell(row=fila_actual+1, column=3, value='5111809000')  # C: Codigo de la cuenta
        ws.cell(row=fila_actual+1, column=4, value='')  # D: Tipo Ident = VACÍO
        ws.cell(row=fila_actual+1, column=5, value='')  # E: No Identificación = VACÍO
        ws.cell(row=fila_actual+1, column=8, value=valor_bruto)  # H: importe (número)
        ws.cell(row=fila_actual+1, column=9, value='WB')  # I: Indicador de IVA
        ws.cell(row=fila_actual+1, column=10, value=rp_doc)  # J: RP Doc Presupuestal ← IMPORTANTE
        ws.cell(row=fila_actual+1, column=11, value=1)  # K: Posc Doc Pres
        ws.cell(row=fila_actual+1, column=26, value=texto_z)  # Z: Texto ← AJUSTE
        
        # ===== FILA P31 (TERCERA FILA DEL BLOQUE) =====
        # Fila 4, 7, 10, 13... - COLUMNA E = no_identificacion
        ws.cell(row=fila_actual+2, column=1, value='P')  # A: Tipo Registro P
        ws.cell(row=fila_actual+2, column=2, value=31)  # B: Clave Contab.
        ws.cell(row=fila_actual+2, column=4, value='CC')  # D: Tipo Ident = "CC"
        ws.cell(row=fila_actual+2, column=5, value=no_identificacion)  # E: No Identificación = DATOS
        ws.cell(row=fila_actual+2, column=7, value='2401010100')  # G: Cuenta contable
        ws.cell(row=fila_actual+2, column=8, value=valor_bruto)  # H: importe (número)
        ws.cell(row=fila_actual+2, column=24, value='0051')  # X: Condicion de Pago
        ws.cell(row=fila_actual+2, column=25, value=asignacion)  # Y: Asignación
        ws.cell(row=fila_actual+2, column=26, value=texto_z)  # Z: Texto ← AJUSTE
        ws.cell(row=fila_actual+2, column=37, value=codigo_bco.zfill(3))  # AM: Código Bco
        ws.cell(row=fila_actual+2, column=38, value=no_cuenta)  # AN: No Cuenta
        ws.cell(row=fila_actual+2, column=39, value=tipo_cta)  # AO: Tipo Cta
        
        # ===== COLUMNAS AN (40) Y AO (41) CON INDICADOR SEGÚN RETECA % =====
        # ESTO ES LO QUE AGREGAMOS SEGÚN TU REQUERIMIENTO
        ws.cell(row=fila_actual+2, column=40, value=indicador_retencion)  # AP: Tipo de retenciones
        ws.cell(row=fila_actual+2, column=41, value=indicador_retencion)  # AQ: Indicador de retención
        
        # COLUMNAS IMPORTANTES CON DATOS REALES:
        ws.cell(row=fila_actual+2, column=42, value=base_retencion)  # AR: Base imponible de retención
        ws.cell(row=fila_actual+2, column=43, value=importe_retencion)  # AS: Importe de retención
        
        print(f"✓ Fila {fila_actual} (C): C='{fecha_actual}', E='1001', F='{fecha_actual}', J='{asignacion}'")
        print(f"✓ Fila {fila_actual+1} (P40): E='', J='{rp_doc}'")
        print(f"✓ Fila {fila_actual+2} (P31): E='{no_identificacion}'")
        print(f"✓ Fila {fila_actual+2} (P31): AN/AO='{indicador_retencion}', AP={base_retencion}, AQ={importe_retencion}")
        
        # Avanzar 3 filas para el siguiente pago
        fila_actual += 3
    
    # Ajustar anchos de columnas
    anchos = {
        'A': 3, 'B': 3, 'C': 12, 'D': 3, 'E': 15, 'F': 12, 'G': 12, 'H': 10,
        'I': 3, 'J': 20, 'K': 25, 'L': 8, 'M': 25, 'N': 8, 'O': 12, 'P': 12,
        'Q': 15, 'R': 8, 'S': 12, 'T': 8, 'U': 15, 'V': 10, 'W': 10, 'X': 15,
        'Y': 12, 'Z': 30, 'AA': 12, 'AB': 20, 'AC': 3, 'AD': 15, 'AE': 10, 'AF': 12,
        'AG': 8, 'AH': 8, 'AI': 8, 'AJ': 15, 'AK': 10, 'AL': 20, 'AM': 8, 'AN': 20,
        'AO': 20, 'AP': 25, 'AQ': 20
    }
    
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho
    
    # Alinear texto a la izquierda
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')
    
    # Guardar archivo
    wb.save(ruta_destino)
    
    # ===== VERIFICAR COLUMNAS IMPORTANTES =====
    print(f"\n{'='*60}")
    print("VERIFICACIÓN DE COLUMNAS CRÍTICAS")
    print('='*60)
    
    print(f"\n📅 Fecha usada en columnas C y F: {fecha_actual}")
    
    print("\nPrimeros 3 bloques (9 filas):")
    print("-" * 80)
    
    for fila in range(2, 11):  # Filas 2 a 10
        valor_c = ws.cell(row=fila, column=3).value  # Columna C (Codigo de la cuenta)
        valor_f = ws.cell(row=fila, column=6).value  # Columna F (Indicador CME)
        valor_e = ws.cell(row=fila, column=5).value  # Columna E
        valor_j = ws.cell(row=fila, column=10).value  # Columna J (RP Doc Presupuestal)
        valor_an = ws.cell(row=fila, column=40).value  # Columna AN (Tipo de retenciones)
        valor_ao = ws.cell(row=fila, column=41).value  # Columna AO (Indicador de retención)
        valor_ap = ws.cell(row=fila, column=42).value  # Columna AP (Base retención)
        valor_aq = ws.cell(row=fila, column=43).value  # Columna AQ (Importe retención)
        tipo = ws.cell(row=fila, column=1).value  # Columna A
        clave = ws.cell(row=fila, column=2).value  # Columna B
        
        if (fila - 2) % 3 == 0:  # Fila C (2, 5, 8)
            tipo_fila = "C"
            c_esperado = fecha_actual
            f_esperado = fecha_actual
            e_esperado = "1001"
            j_esperado = "Asignación"
        elif (fila - 2) % 3 == 1:  # Fila P40 (3, 6, 9)
            tipo_fila = "P40"
            c_esperado = "5111809000"
            f_esperado = "VACÍO"
            e_esperado = "VACÍO"
            j_esperado = "RP Doc"
        else:  # Fila P31 (4, 7, 10)
            tipo_fila = "P31"
            c_esperado = "VACÍO"
            f_esperado = "VACÍO"
            e_esperado = "DATOS"
            j_esperado = "VACÍO"
        
        print(f"Fila {fila:2d} ({tipo_fila}): A='{tipo}', B={clave}")
        print(f"  Col C: '{valor_c}' (Esperado: {c_esperado})")
        print(f"  Col F: '{valor_f}' (Esperado: {f_esperado})")
        print(f"  Col E: '{valor_e}' (Esperado: {e_esperado})")
        print(f"  Col J (RP Doc): '{valor_j}' (Esperado: {j_esperado})")
        
        if tipo_fila == "P31":
            print(f"  Col AN (Tipo ret): '{valor_an}'")
            print(f"  Col AO (Ind ret): '{valor_ao}'")
            print(f"  Col AP (Base): {valor_ap}")
            print(f"  Col AQ (Importe): {valor_aq}")
        print()
    
    # Estadísticas
    print(f"\n{'='*60}")
    print("ESTADÍSTICAS DE DATOS")
    print('='*60)
    
    total_c = 0
    c_correctas = 0
    total_p40 = 0
    j_con_datos = 0
    total_p31 = 0
    an_con_datos = 0
    ao_con_datos = 0
    ap_con_datos = 0
    aq_con_datos = 0
    
    for fila in range(2, fila_actual):
        tipo = ws.cell(row=fila, column=1).value
        clave = ws.cell(row=fila, column=2).value
        
        if tipo == 'C':
            total_c += 1
            if ws.cell(row=fila, column=3).value == fecha_actual:  # Columna C
                c_correctas += 1
        
        if tipo == 'P' and clave == 40:
            total_p40 += 1
            if ws.cell(row=fila, column=10).value not in [None, '', ' ']:  # Columna J
                j_con_datos += 1
        
        if tipo == 'P' and clave == 31:
            total_p31 += 1
            if ws.cell(row=fila, column=40).value not in [None, '', ' ']:  # Columna AN
                an_con_datos += 1
            if ws.cell(row=fila, column=41).value not in [None, '', ' ']:  # Columna AO
                ao_con_datos += 1
            if ws.cell(row=fila, column=42).value not in [None, 0, '']:  # Columna AP
                ap_con_datos += 1
            if ws.cell(row=fila, column=43).value not in [None, 0, '']:  # Columna AQ
                aq_con_datos += 1
    
    print(f"Total filas C: {total_c}")
    print(f"Filas C con fecha actual en columna C: {c_correctas} ({c_correctas/total_c*100:.1f}%)")
    print(f"\nTotal filas P40: {total_p40}")
    print(f"Filas P40 con datos en J (RP Doc): {j_con_datos} ({j_con_datos/total_p40*100:.1f}%)")
    print(f"\nTotal filas P31: {total_p31}")
    print(f"Filas P31 con datos en AN (Tipo ret): {an_con_datos} ({an_con_datos/total_p31*100:.1f}%)")
    print(f"Filas P31 con datos en AO (Ind ret): {ao_con_datos} ({ao_con_datos/total_p31*100:.1f}%)")
    print(f"Filas P31 con datos en AP (Base): {ap_con_datos} ({ap_con_datos/total_p31*100:.1f}%)")
    print(f"Filas P31 con datos en AQ (Importe): {aq_con_datos} ({aq_con_datos/total_p31*100:.1f}%)")
    
    # Estadísticas de indicadores usados
    indicadores_usados = {}
    for fila in range(2, fila_actual):
        if ws.cell(row=fila, column=1).value == 'P' and ws.cell(row=fila, column=2).value == 31:
            indicador = ws.cell(row=fila, column=40).value
            if indicador:
                indicadores_usados[indicador] = indicadores_usados.get(indicador, 0) + 1
    
    if indicadores_usados:
        print(f"\n📊 Indicadores de retención usados:")
        for indicador, count in sorted(indicadores_usados.items()):
            print(f"  {indicador}: {count} veces")
    
    print(f"\n{'='*60}")
    print(f"¡ARCHIVO GENERADO EXITOSAMENTE!")
    print(f"Ubicación: {ruta_destino}")
    print(f"Total de pagos procesados: {len(df)}")
    print(f"Total de filas generadas: {fila_actual - 1}")
    print('='*60)
    
    return True

# Ejecutar
if __name__ == "__main__":
    print("="*60)
    print("GENERADOR DE PLANTILLA DE PAGOS - CON FECHA ACTUAL Y RETECA %")
    print("="*60)
    
    if procesar_pagos_consolidado():
        print("\n✓ ARCHIVO GENERADO CON ÉXITO")
        print("\nVERIFICA EN EL ARCHIVO GENERADO:")
        print(f"1. Columnas C y F con fecha actual ({datetime.now().strftime('%Y%m%d')}):")
        print("   • C2, C5, C8, C11... = fecha actual")
        print("   • F2, F5, F8, F11... = fecha actual")
        print("\n2. Columna E:")
        print("   • E2, E5, E8, E11... = '1001'")
        print("   • E3, E6, E9, E12... = vacío")
        print("   • E4, E7, E10, E13... = datos del consolidado")
        print("\n3. Columna J (RP Doc Presupuestal):")
        print("   • J2, J5, J8, J11... = asignación (ej: 054-2025)")
        print("   • J3, J6, J9, J12... = RP Doc del consolidado")
        print("   • J4, J7, J10, J13... = vacío")
        print("\n4. Columnas AN y AO (RETECA % → INDICADOR):")
        print("   • AN (Tipo de retenciones) = indicador según Reteica %")
        print("   • AO (Indicador de retención) = mismo indicador")
        print("   Ej: 0,966% → '39', 0,100% → '01', etc.")
        print("\n5. Columnas AP y AQ (retenciones):")
        print("   • AP (Base imponible de retención)")
        print("   • AQ (Importe de retención)")
        print("   Ambas deben tener datos reales del consolidado")
    else:
        print("✗ Error en el proceso")