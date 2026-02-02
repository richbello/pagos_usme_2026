import pdfplumber
import pandas as pd
import os
import re
from openpyxl import load_workbook
from datetime import datetime

# --- CONFIGURACIÓN ---
ruta_carpeta = r'C:\RICHARD\FDL\Usme\2026\Pruebas_pagos'
ruta_plantilla = os.path.join(ruta_carpeta, "FORMATO FIJO PAGOS.xlsx")
archivo_salida = os.path.join(ruta_carpeta, "CARGUE_MASIVO_CONSOLIDADO.xlsx")

def limpiar_a_entero(texto):
    if not texto: return 0
    limpio = re.sub(r'[^\d]', '', str(texto))
    try: return int(limpio)
    except: return 0

def procesar_pdf(pdf_path):
    datos = {
        'Contratista': "N/A", 'Contrato': "N/A", 'Contrato_Corto': "N/A", 
        'NIT_CC': "", 'Bruto': 0, 'Descuentos': 0
    }
    with pdfplumber.open(pdf_path) as pdf:
        pagina = pdf.pages[0]
        texto = pagina.extract_text()
        tablas = pagina.extract_tables()
        
        nom_m = re.search(r'CONTRATISTA:\s*([A-Z\sÁÉÍÓÚÑ]+)', texto)
        if nom_m: datos['Contratista'] = nom_m.group(1).strip()
        
        id_m = re.search(r'\d{1,3}(?:\.\d{3}){2,3}', texto)
        if id_m: datos['NIT_CC'] = id_m.group(0).replace('.', '')
        
        cnt = re.search(r'CPS[-\s]*(\d+[-\s]*\d+)', texto)
        if cnt:
            datos['Contrato'] = f"CPS-{cnt.group(1)}"
            datos['Contrato_Corto'] = cnt.group(1)

        for tabla in tablas:
            for fila in tabla:
                linea = " ".join([str(c) for c in fila if c]).upper()
                if "VALOR BRUTO" in linea: 
                    datos['Bruto'] = limpiar_a_entero(fila[-1])
                if "TOTAL DESCUENTOS" in linea:
                    datos['Descuentos'] = limpiar_a_entero(fila[-1])
    return datos

def generar_consolidado():
    wb = load_workbook(ruta_plantilla)
    ws = wb.active
    while ws.max_row > 1: ws.delete_rows(2)

    archivos = [f for f in os.listdir(ruta_carpeta) if f.lower().endswith(".pdf") and not f.startswith("~$")]
    fecha_hoy = datetime.now().strftime("%Y%m%d")
    
    fila_actual = 2
    contador = 1

    for pdf in archivos:
        info = procesar_pdf(os.path.join(ruta_carpeta, pdf))
        
        # --- FILA C (Cabecera) ---
        ws.cell(row=fila_actual, column=1).value = "C"
        ws.cell(row=fila_actual, column=2).value = contador
        ws.cell(row=fila_actual, column=3).value = fecha_hoy
        ws.cell(row=fila_actual, column=4).value = "KR"
        ws.cell(row=fila_actual, column=6).value = fecha_hoy
        ws.cell(row=fila_actual, column=8).value = "COP"
        ws.cell(row=fila_actual, column=10).value = info['Contrato']
        ws.cell(row=fila_actual, column=11).value = info['Contratista']
        fila_actual += 1

        # --- FILA P1 (Gasto) ---
        ws.cell(row=fila_actual, column=1).value = "P"
        ws.cell(row=fila_actual, column=2).value = "40"
        ws.cell(row=fila_actual, column=3).value = "5111809000"
        ws.cell(row=fila_actual, column=8).value = info['Bruto'] # H3, H6...
        ws.cell(row=fila_actual, column=9).value = "WB"
        ws.cell(row=fila_actual, column=11).value = 1
        fila_actual += 1

        # --- FILA P2 (Tercero) ---
        ws.cell(row=fila_actual, column=1).value = "P"
        ws.cell(row=fila_actual, column=2).value = "31"
        ws.cell(row=fila_actual, column=4).value = "CC"
        ws.cell(row=fila_actual, column=5).value = info['NIT_CC']
        ws.cell(row=fila_actual, column=7).value = "2401010100"
        
        # Valor Bruto en H4, H7, H10...
        valor_bruto = info['Bruto']
        ws.cell(row=fila_actual, column=8).value = valor_bruto 
        
        ws.cell(row=fila_actual, column=24).value = "0051"
        ws.cell(row=fila_actual, column=25).value = info['Contrato_Corto']
        ws.cell(row=fila_actual, column=26).value = f"PAGO {info['Contrato']}"
        
        # --- CÁLCULO MATEMÁTICO SOLICITADO ---
        # AP = H * 0.886 (Sin decimales)
        base_reteica_calculada = int(valor_bruto * 0.886)
        
        ws.cell(row=fila_actual, column=42).value = base_reteica_calculada # AP
        ws.cell(row=fila_actual, column=43).value = info['Descuentos']      # AQ
        
        fila_actual += 1
        contador += 1

    wb.save(archivo_salida)
    print(f"Consolidado generado. AP calculado como Bruto * 0.886")

if __name__ == "__main__":
    generar_consolidado()