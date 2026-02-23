import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ============================================================
# RUTAS
# ============================================================
ruta_pagos  = r"C:\RICHARD\FDL\Usme\2026\Pagos\Febrero\PLANTILLA_PAGOS_GENERADAFEB22.xlsx"
ruta_crp    = r"C:\RICHARD\FDL\Usme\2026\Reporte CRP\Febrero\Reporte CRP Feb22.xlsx"
ruta_pac    = r"C:\RICHARD\FDL\Usme\2026\Pac\Febrero\Reporte PAC Feb23_130pm.xlsx"
ruta_salida = r"C:\RICHARD\FDL\Usme\2026\Pagos\Febrero\RESULTADO_FEB22.xlsx"

# ============================================================
# MES ACTUAL → período como string "02", "03", etc.
# ============================================================
mes_actual = str(date.today().month).zfill(2)
MESES = {
    '01':'ENERO','02':'FEBRERO','03':'MARZO','04':'ABRIL',
    '05':'MAYO','06':'JUNIO','07':'JULIO','08':'AGOSTO',
    '09':'SEPTIEMBRE','10':'OCTUBRE','11':'NOVIEMBRE','12':'DICIEMBRE'
}

# ============================================================
# 1. LEER PLANTILLA PAGOS
# ============================================================
df_pagos = pd.read_excel(ruta_pagos, sheet_name=0, header=0, dtype=str)
df_pagos.columns = df_pagos.columns.str.strip()

df_pagos_p = df_pagos[
    (df_pagos['Tipo Registro P'].str.strip() == 'P') &
    (df_pagos['Clave Contab.'].str.strip() == '40')
].copy()

df_pagos_p['CRP']     = df_pagos_p['RP Doc Presupuestal'].astype(str).str.strip()
df_pagos_p['IMPORTE'] = pd.to_numeric(df_pagos_p['importe'], errors='coerce').fillna(0)
df_pagos_p = df_pagos_p[['CRP', 'IMPORTE']].dropna(subset=['CRP'])
df_pagos_p = df_pagos_p[df_pagos_p['CRP'] != 'nan']

# ============================================================
# 2. LEER REPORTE CRP
# ============================================================
df_crp = pd.read_excel(ruta_crp, sheet_name=0, header=0, dtype=str)
df_crp.columns = df_crp.columns.str.strip()
df_crp['crp']    = df_crp['crp'].astype(str).str.strip()
df_crp['Rubro']  = df_crp['Rubro'].astype(str).str.strip()
df_crp['Fondos'] = df_crp['Fondos'].astype(str).str.strip()
df_crp_uniq = df_crp[['crp', 'Rubro', 'Fondos']].drop_duplicates(subset='crp')

# ============================================================
# 3. CRUZAR PAGOS + CRP
# ============================================================
df_merged = df_pagos_p.merge(
    df_crp_uniq, left_on='CRP', right_on='crp', how='left'
).drop(columns=['crp'])
df_merged.rename(columns={'IMPORTE': 'VR A PAGAR'}, inplace=True)
df_merged['Rubro']  = df_merged['Rubro'].fillna('SIN RUBRO')
df_merged['Fondos'] = df_merged['Fondos'].fillna('SIN FONDO')

# Totales por Rubro + Fondo
totales_rubro = df_merged.groupby(['Rubro', 'Fondos'])['VR A PAGAR'].sum().reset_index()
totales_rubro.rename(columns={'VR A PAGAR': 'TOTAL_A_PAGAR'}, inplace=True)

# ============================================================
# 4. LEER REPORTE PAC
# ============================================================
df_pac = pd.read_excel(ruta_pac, sheet_name='Data', header=0, dtype=str)
df_pac.columns = df_pac.columns.str.strip()

df_pac['Per.presup.']       = df_pac['Per.presup.'].astype(str).str.strip().str.zfill(2)
df_pac['Pos.presupuestaria']= df_pac['Pos.presupuestaria'].astype(str).str.strip()
df_pac['Fondos']            = df_pac['Fondos'].astype(str).str.strip()
df_pac['Disponibilidad PAC']= pd.to_numeric(df_pac['Disponibilidad PAC'], errors='coerce').fillna(0)

# Solo filas con datos reales (excluir subtotales/encabezados sin período válido)
df_pac_data = df_pac[df_pac['Per.presup.'].str.match(r'^\d{2}$')].copy()
df_pac_data = df_pac_data[df_pac_data['Pos.presupuestaria'] != 'nan']
df_pac_data = df_pac_data[df_pac_data['Disponibilidad PAC'] > 0]

# ============================================================
# 5. ANÁLISIS PAC POR RUBRO / FONDO
# ============================================================
alertas = []

for _, row in totales_rubro.iterrows():
    rubro        = row['Rubro']
    fondo        = row['Fondos']
    total_pagar  = row['TOTAL_A_PAGAR']

    # Filtrar PAC para este rubro y fondo
    mask = (
        (df_pac_data['Pos.presupuestaria'] == rubro) &
        (df_pac_data['Fondos'] == fondo)
    )
    df_rubro_pac = df_pac_data[mask].copy()

    # --- MES ACTUAL ---
    pac_mes_actual = df_rubro_pac[df_rubro_pac['Per.presup.'] == mes_actual]['Disponibilidad PAC'].sum()

    if pac_mes_actual >= total_pagar:
        estado = '✅ ALCANZA'
        detalle = (
            f"  Rubro: {rubro} | Fondo: {fondo}\n"
            f"  Total a pagar : $ {total_pagar:>20,.0f}\n"
            f"  Disp. PAC {MESES.get(mes_actual,'MES ACTUAL'):10}: $ {pac_mes_actual:>20,.0f}\n"
            f"  → {estado}: La disponibilidad PAC del mes cubre el pago.\n"
        )
    else:
        estado = '❌ NO ALCANZA'
        detalle = (
            f"  Rubro: {rubro} | Fondo: {fondo}\n"
            f"  Total a pagar       : $ {total_pagar:>20,.0f}\n"
            f"  Disp. PAC {MESES.get(mes_actual,'MES ACTUAL'):10}: $ {pac_mes_actual:>20,.0f}\n"
            f"  Diferencia faltante : $ {(total_pagar - pac_mes_actual):>20,.0f}\n"
            f"  → {estado} en el mes actual. Recursos disponibles en otros meses:\n"
        )
        # Buscar otros meses con disponibilidad
        otros_meses = df_rubro_pac[df_rubro_pac['Per.presup.'] != mes_actual].sort_values('Per.presup.')
        if not otros_meses.empty:
            for _, m in otros_meses.iterrows():
                nombre_mes = MESES.get(m['Per.presup.'], m['Per.presup.'])
                detalle += f"      Período {m['Per.presup.']} - {nombre_mes:<12}: $ {m['Disponibilidad PAC']:>15,.0f}\n"
            total_otros = otros_meses['Disponibilidad PAC'].sum()
            detalle += f"      {'TOTAL OTROS MESES':<20}: $ {total_otros:>15,.0f}\n"
            if (pac_mes_actual + total_otros) >= total_pagar:
                detalle += f"      ⚠️  Sumando otros meses SÍ se puede cubrir el pago.\n"
            else:
                detalle += f"      🚨 ADVERTENCIA: Incluso sumando todos los meses NO alcanza el PAC.\n"
        else:
            detalle += "      🚨 NO hay disponibilidad PAC en ningún otro mes para este rubro/fondo.\n"

    alertas.append({'rubro': rubro, 'fondo': fondo, 'estado': estado, 'detalle': detalle})

# ============================================================
# 6. IMPRIMIR ALERTAS EN CONSOLA
# ============================================================
print("\n" + "="*65)
print("           ANÁLISIS DISPONIBILIDAD PAC vs PAGOS")
print("="*65)
for a in alertas:
    print(f"\n{a['estado']} — Rubro: {a['rubro']} | Fondo: {a['fondo']}")
    print(a['detalle'])

# ============================================================
# 7. CONSTRUIR TABLA RESULTADO EXCEL
# ============================================================
filas_resultado = []

for rubro, grupo in df_merged.groupby('Rubro', sort=True):
    fondos_val  = grupo['Fondos'].iloc[0]
    for _, row in grupo.iterrows():
        filas_resultado.append({
            'CRP': row['CRP'], 'VR A PAGAR': row['VR A PAGAR'],
            'RUBRO': row['Rubro'], 'FONDOS': fondos_val, 'TIPO': 'dato'
        })
    subtotal = grupo['VR A PAGAR'].sum()
    filas_resultado.append({'CRP':'','VR A PAGAR': subtotal,'RUBRO':'','FONDOS':'','TIPO':'subtotal'})
    filas_resultado.append({'CRP':'','VR A PAGAR':'','RUBRO':'','FONDOS':'','TIPO':'vacio'})

total_general = df_merged['VR A PAGAR'].sum()
filas_resultado.append({'CRP':'TOTAL GENERAL','VR A PAGAR': total_general,'RUBRO':'','FONDOS':'','TIPO':'total'})

# Agregar sección PAC al Excel
filas_resultado.append({'CRP':'','VR A PAGAR':'','RUBRO':'','FONDOS':'','TIPO':'vacio'})
filas_resultado.append({'CRP':'═══ ANÁLISIS DISPONIBILIDAD PAC ═══','VR A PAGAR':'','RUBRO':'','FONDOS':'','TIPO':'header_pac'})

for a in alertas:
    # Fila estado
    filas_resultado.append({
        'CRP': a['estado'],
        'VR A PAGAR': '',
        'RUBRO': a['rubro'],
        'FONDOS': a['fondo'],
        'TIPO': 'alerta_ok' if '✅' in a['estado'] else 'alerta_no'
    })
    # PAC mes actual
    mask = (
        (df_pac_data['Pos.presupuestaria'] == a['rubro']) &
        (df_pac_data['Fondos'] == a['fondo']) &
        (df_pac_data['Per.presup.'] == mes_actual)
    )
    pac_actual = df_pac_data[mask]['Disponibilidad PAC'].sum()
    total_pg   = totales_rubro[(totales_rubro['Rubro']==a['rubro']) & (totales_rubro['Fondos']==a['fondo'])]['TOTAL_A_PAGAR'].values[0]

    filas_resultado.append({
        'CRP': f"  Total a pagar", 'VR A PAGAR': total_pg,
        'RUBRO': f"  Disp. PAC {MESES.get(mes_actual,'')}", 'FONDOS': pac_actual, 'TIPO': 'pac_detalle'
    })

    # Si no alcanza, otros meses
    if '❌' in a['estado']:
        mask2 = (
            (df_pac_data['Pos.presupuestaria'] == a['rubro']) &
            (df_pac_data['Fondos'] == a['fondo']) &
            (df_pac_data['Per.presup.'] != mes_actual)
        )
        otros = df_pac_data[mask2].sort_values('Per.presup.')
        for _, m in otros.iterrows():
            nombre_mes = MESES.get(m['Per.presup.'], m['Per.presup.'])
            filas_resultado.append({
                'CRP': f"  Período {m['Per.presup.']} - {nombre_mes}",
                'VR A PAGAR': '', 'RUBRO': '  Disp. PAC',
                'FONDOS': m['Disponibilidad PAC'], 'TIPO': 'pac_otro_mes'
            })
    filas_resultado.append({'CRP':'','VR A PAGAR':'','RUBRO':'','FONDOS':'','TIPO':'vacio'})

df_resultado = pd.DataFrame(filas_resultado)
df_resultado[['CRP','VR A PAGAR','RUBRO','FONDOS']].to_excel(ruta_salida, index=False, sheet_name='RESULTADO')

# ============================================================
# 8. FORMATO EXCEL
# ============================================================
wb = load_workbook(ruta_salida)
ws = wb['RESULTADO']

header_fill    = PatternFill("solid", fgColor="1F4E79")
subtotal_fill  = PatternFill("solid", fgColor="BDD7EE")
total_fill     = PatternFill("solid", fgColor="F4B942")
ok_fill        = PatternFill("solid", fgColor="C6EFCE")
no_fill        = PatternFill("solid", fgColor="FFC7CE")
pac_fill       = PatternFill("solid", fgColor="EBF3FB")
otro_mes_fill  = PatternFill("solid", fgColor="FFF2CC")
header_pac_fill= PatternFill("solid", fgColor="2E4057")

header_font    = Font(bold=True, color="FFFFFF", size=11)
bold_font      = Font(bold=True, size=10)
normal_font    = Font(size=10)
white_bold     = Font(bold=True, color="FFFFFF", size=11)

thin   = Side(style='thin', color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, w in enumerate([18, 20, 20, 18], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Encabezado fila 1
for cell in ws[1]:
    cell.fill = header_fill; cell.font = header_font
    cell.alignment = Alignment(horizontal="center"); cell.border = border

tipo_list = [f['TIPO'] for f in filas_resultado]

for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 0):
    tipo = tipo_list[idx] if idx < len(tipo_list) else 'dato'
    crp_val = str(row[0].value or '')

    for cell in row:
        cell.border = border; cell.font = normal_font

    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = '"$ "#,##0'
            cell.alignment = Alignment(horizontal="right")

    if tipo == 'subtotal':
        for cell in row: cell.fill = subtotal_fill; cell.font = bold_font
    elif tipo == 'total':
        for cell in row: cell.fill = total_fill; cell.font = bold_font
    elif tipo == 'header_pac':
        for cell in row: cell.fill = header_pac_fill; cell.font = white_bold
    elif tipo == 'alerta_ok':
        for cell in row: cell.fill = ok_fill; cell.font = bold_font
    elif tipo == 'alerta_no':
        for cell in row: cell.fill = no_fill; cell.font = bold_font
    elif tipo == 'pac_detalle':
        for cell in row: cell.fill = pac_fill
    elif tipo == 'pac_otro_mes':
        for cell in row: cell.fill = otro_mes_fill

ws.freeze_panes = "A2"
wb.save(ruta_salida)

print(f"\n✅ Archivo generado: {ruta_salida}")
print(f"   Total general a pagar: $ {total_general:,.0f}")